"""
Report Service for generating various reports
"""

import io
from uuid import UUID
import os
from typing import Dict, Any, List, Optional
from datetime import date, datetime, timedelta
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func
from backend.models.daily_log import DailyLog
from backend.models.spot import Spot, SpotStatus
from backend.models.copy import Copy
from backend.models.order import Order, OrderStatus
from backend.models.invoice import Invoice
from backend.models.makegood import Makegood
from backend.models.advertiser import Advertiser
from backend.models.sales_rep import SalesRep
from backend.models.audio_cut import AudioCut
from backend.models.live_read import LiveRead
from backend.models.political_record import PoliticalRecord
from backend.models.audio_delivery import AudioDelivery
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import structlog

logger = structlog.get_logger()


class ReportService:
    """Service for generating reports"""
    
    def __init__(self, db: AsyncSession):
        self.db = db
    
    async def generate_daily_log_report(self, log_date: date, station_id: Optional[UUID] = None) -> Dict[str, Any]:
        """Generate daily log report"""
        query = select(DailyLog).where(DailyLog.date == log_date)
        
        if station_id is not None:
            query = query.where(DailyLog.station_id == station_id)
        
        result = await self.db.execute(query)
        log = result.scalar_one_or_none()
        
        if not log:
            return {"date": log_date.isoformat(), "log_exists": False, "station_id": station_id}
        
        return {
            "date": log_date.isoformat(),
            "log_id": log.id,
            "published": log.published,
            "locked": log.locked,
            "total_elements": sum(
                len(hour_data.get("elements", []))
                for hour_data in log.json_data.get("hourly_logs", {}).values()
            ),
            "conflicts": log.conflicts or [],
            "oversell_warnings": log.oversell_warnings or []
        }
    
    async def generate_missing_copy_report(
        self,
        start_date: Optional[date],
        end_date: Optional[date],
        station_id: Optional[UUID] = None
    ) -> Dict[str, Any]:
        """Generate missing copy report"""
        # Get spots without copy assignments
        query = select(Spot).where(Spot.status == SpotStatus.SCHEDULED)
        
        if start_date:
            query = query.where(Spot.scheduled_date >= start_date)
        if end_date:
            query = query.where(Spot.scheduled_date <= end_date)
        if station_id is not None:
            query = query.where(Spot.station_id == station_id)
        
        result = await self.db.execute(query)
        spots = result.scalars().all()
        
        # Check which spots have copy
        from backend.models.copy_assignment import CopyAssignment
        
        missing_copy = []
        for spot in spots:
            assignment_result = await self.db.execute(
                select(CopyAssignment).where(CopyAssignment.spot_id == spot.id)
            )
            if not assignment_result.scalar_one_or_none():
                missing_copy.append({
                    "spot_id": spot.id,
                    "order_id": spot.order_id,
                    "scheduled_date": spot.scheduled_date.isoformat(),
                    "scheduled_time": spot.scheduled_time
                })
        
        return {
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
            "missing_copy_count": len(missing_copy),
            "spots_missing_copy": missing_copy
        }
    
    async def generate_avails_report(
        self,
        start_date: date,
        end_date: date
    ) -> Dict[str, Any]:
        """Generate avails report"""
        from backend.services.spot_scheduler import SpotScheduler
        
        scheduler = SpotScheduler(self.db)
        
        avails_by_date = {}
        current_date = start_date
        while current_date <= end_date:
            avails = await scheduler.calculate_avails(current_date)
            avails_by_date[current_date.isoformat()] = avails
            current_date += timedelta(days=1)
        
        return {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "avails_by_date": avails_by_date
        }
    
    async def generate_conflicts_report(self, log_date: date) -> Dict[str, Any]:
        """Generate conflict report"""
        from backend.services.spot_scheduler import SpotScheduler
        
        scheduler = SpotScheduler(self.db)
        conflicts = await scheduler.detect_conflicts(log_date)
        
        return {
            "date": log_date.isoformat(),
            "conflict_count": len(conflicts),
            "conflicts": conflicts
        }
    
    async def generate_expirations_report(self, days_ahead: int) -> Dict[str, Any]:
        """Generate expiring copy/contracts report"""
        from backend.services.copy_service import CopyService
        
        copy_service = CopyService(self.db)
        expiring_copy = await copy_service.check_expiring(days_ahead)
        
        # Get expiring orders
        cutoff_date = date.today() + timedelta(days=days_ahead)
        result = await self.db.execute(
            select(Order).where(
                and_(
                    Order.end_date <= cutoff_date,
                    Order.end_date >= date.today(),
                    Order.status == OrderStatus.ACTIVE
                )
            )
        )
        expiring_orders = result.scalars().all()
        
        return {
            "days_ahead": days_ahead,
            "expiring_copy": [
                {
                    "id": c.id,
                    "title": c.title,
                    "expires_at": c.expires_at.isoformat() if c.expires_at else None
                }
                for c in expiring_copy
            ],
            "expiring_orders": [
                {
                    "id": o.id,
                    "order_number": o.order_number,
                    "end_date": o.end_date.isoformat()
                }
                for o in expiring_orders
            ]
        }
    
    async def generate_revenue_summary(
        self,
        start_date: date,
        end_date: date
    ) -> Dict[str, Any]:
        """Generate revenue summary"""
        result = await self.db.execute(
            select(func.sum(Invoice.total)).where(
                and_(
                    Invoice.invoice_date >= start_date,
                    Invoice.invoice_date <= end_date,
                    Invoice.status.in_(["SENT", "PAID"])
                )
            )
        )
        total_revenue = result.scalar() or 0
        
        result = await self.db.execute(
            select(func.count(Invoice.id)).where(
                and_(
                    Invoice.invoice_date >= start_date,
                    Invoice.invoice_date <= end_date
                )
            )
        )
        invoice_count = result.scalar() or 0
        
        return {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "total_revenue": float(total_revenue),
            "invoice_count": invoice_count
        }
    
    async def generate_makegood_report(
        self,
        start_date: Optional[date],
        end_date: Optional[date]
    ) -> Dict[str, Any]:
        """Generate makegood report"""
        query = select(Makegood)
        
        if start_date:
            query = query.where(Makegood.created_at >= datetime.combine(start_date, datetime.min.time()))
        if end_date:
            query = query.where(Makegood.created_at <= datetime.combine(end_date, datetime.max.time()))
        
        result = await self.db.execute(query)
        makegoods = result.scalars().all()
        
        return {
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
            "makegood_count": len(makegoods),
            "makegoods": [
                {
                    "id": mg.id,
                    "original_spot_id": mg.original_spot_id,
                    "makegood_spot_id": mg.makegood_spot_id,
                    "reason": mg.reason,
                    "approved": mg.approved_by is not None
                }
                for mg in makegoods
            ]
        }
    
    async def generate_revenue_by_rep(
        self,
        start_date: date,
        end_date: date
    ) -> Dict[str, Any]:
        """Generate revenue by rep report"""
        result = await self.db.execute(
            select(
                SalesRep.id,
                func.sum(Invoice.total).label("total_revenue")
            ).join(Order, Order.sales_rep_id == SalesRep.id)
            .join(Invoice, Invoice.order_id == Order.id)
            .where(
                and_(
                    Invoice.invoice_date >= start_date,
                    Invoice.invoice_date <= end_date
                )
            )
            .group_by(SalesRep.id)
        )
        
        revenue_by_rep = []
        for row in result.all():
            revenue_by_rep.append({
                "sales_rep_id": row.id,
                "total_revenue": float(row.total_revenue)
            })
        
        return {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "revenue_by_rep": revenue_by_rep
        }
    
    async def generate_revenue_by_advertiser(
        self,
        start_date: date,
        end_date: date
    ) -> Dict[str, Any]:
        """Generate revenue by advertiser report"""
        result = await self.db.execute(
            select(
                Advertiser.id,
                Advertiser.name,
                func.sum(Invoice.total).label("total_revenue")
            ).join(Invoice, Invoice.advertiser_id == Advertiser.id)
            .where(
                and_(
                    Invoice.invoice_date >= start_date,
                    Invoice.invoice_date <= end_date
                )
            )
            .group_by(Advertiser.id, Advertiser.name)
        )
        
        revenue_by_advertiser = []
        for row in result.all():
            revenue_by_advertiser.append({
                "advertiser_id": row.id,
                "advertiser_name": row.name,
                "total_revenue": float(row.total_revenue)
            })
        
        return {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "revenue_by_advertiser": revenue_by_advertiser
        }
    
    async def generate_pending_orders_report(self) -> Dict[str, Any]:
        """Generate pending orders report"""
        result = await self.db.execute(
            select(Order).where(
                Order.status.in_([OrderStatus.PENDING, OrderStatus.DRAFT])
            )
        )
        orders = result.scalars().all()
        
        return {
            "pending_orders": [
                {
                    "id": o.id,
                    "order_number": o.order_number,
                    "advertiser_id": o.advertiser_id,
                    "status": o.status.value,
                    "total_value": float(o.total_value)
                }
                for o in orders
            ],
            "count": len(orders)
        }
    
    async def generate_expiring_contracts_report(self, days_ahead: int) -> Dict[str, Any]:
        """Generate expiring contracts report"""
        cutoff_date = date.today() + timedelta(days=days_ahead)
        
        result = await self.db.execute(
            select(Order).where(
                and_(
                    Order.end_date <= cutoff_date,
                    Order.end_date >= date.today(),
                    Order.status == OrderStatus.ACTIVE
                )
            )
        )
        orders = result.scalars().all()
        
        return {
            "days_ahead": days_ahead,
            "expiring_contracts": [
                {
                    "id": o.id,
                    "order_number": o.order_number,
                    "advertiser_id": o.advertiser_id,
                    "end_date": o.end_date.isoformat()
                }
                for o in orders
            ],
            "count": len(orders)
        }
    
    async def generate_audio_activity_report(
        self,
        start_date: date,
        end_date: date
    ) -> Dict[str, Any]:
        """Generate audio activity report showing all cuts delivered vs scheduled"""
        # Get all cuts with delivery records in date range
        result = await self.db.execute(
            select(AudioDelivery).where(
                and_(
                    AudioDelivery.created_at >= datetime.combine(start_date, datetime.min.time()),
                    AudioDelivery.created_at <= datetime.combine(end_date, datetime.max.time())
                )
            )
        )
        deliveries = result.scalars().all()
        
        report_data = []
        for delivery in deliveries:
            cut_result = await self.db.execute(
                select(AudioCut).where(AudioCut.id == delivery.cut_id)
            )
            cut = cut_result.scalar_one_or_none()
            
            if cut:
                report_data.append({
                    "cut_id": cut.cut_id,
                    "cut_name": cut.cut_name,
                    "copy_id": cut.copy_id,
                    "delivery_method": delivery.delivery_method,
                    "target_server": delivery.target_server,
                    "status": delivery.status,
                    "delivered_at": delivery.delivery_completed_at.isoformat() if delivery.delivery_completed_at else None,
                    "checksum_verified": delivery.checksum_verified
                })
        
        return {
            "report_type": "audio_activity",
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "data": report_data,
            "total": len(report_data)
        }
    
    async def generate_cut_rotation_performance(
        self,
        copy_id: Optional[UUID] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> Dict[str, Any]:
        """Generate cut rotation performance report"""
        query = select(AudioCut)
        
        if copy_id:
            query = query.where(AudioCut.copy_id == copy_id)
        
        result = await self.db.execute(query)
        cuts = result.scalars().all()
        
        report_data = []
        for cut in cuts:
            # Get delivery count for this cut
            delivery_count_result = await self.db.execute(
                select(func.count(AudioDelivery.id)).where(
                    and_(
                        AudioDelivery.cut_id == cut.id,
                        AudioDelivery.status == "success"
                    )
                )
            )
            delivery_count = delivery_count_result.scalar() or 0
            
            report_data.append({
                "cut_id": cut.cut_id,
                "cut_name": cut.cut_name,
                "copy_id": cut.copy_id,
                "rotation_weight": cut.rotation_weight,
                "active": cut.active,
                "delivery_count": delivery_count,
                "version": cut.version
            })
        
        return {
            "report_type": "cut_rotation_performance",
            "copy_id": copy_id,
            "data": report_data,
            "total": len(report_data)
        }
    
    async def generate_live_read_performance(
        self,
        start_date: date,
        end_date: date
    ) -> Dict[str, Any]:
        """Generate live read performance report"""
        result = await self.db.execute(
            select(LiveRead).where(
                and_(
                    LiveRead.scheduled_time >= datetime.combine(start_date, datetime.min.time()),
                    LiveRead.scheduled_time <= datetime.combine(end_date, datetime.max.time())
                )
            )
        )
        live_reads = result.scalars().all()
        
        report_data = []
        for read in live_reads:
            report_data.append({
                "id": read.id,
                "script_title": read.script_title,
                "scheduled_time": read.scheduled_time.isoformat() if read.scheduled_time else None,
                "performed_time": read.performed_time.isoformat() if read.performed_time else None,
                "status": read.status,
                "confirmed": read.confirmed,
                "makegood_required": read.makegood_required
            })
        
        return {
            "report_type": "live_read_performance",
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "data": report_data,
            "total": len(report_data)
        }
    
    async def generate_fcc_compliance_log(
        self,
        start_date: date,
        end_date: date
    ) -> Dict[str, Any]:
        """Generate FCC compliance log for political ads"""
        from backend.services.political_compliance_service import PoliticalComplianceService
        
        service = PoliticalComplianceService(self.db)
        compliance_log = await service.generate_fcc_compliance_log(
            datetime.combine(start_date, datetime.min.time()),
            datetime.combine(end_date, datetime.max.time())
        )
        
        return {
            "report_type": "fcc_compliance",
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "data": compliance_log,
            "total": len(compliance_log)
        }
    
    async def export_report(
        self,
        report_type: str,
        format: str,
        report_params: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Export report to PDF/Excel/CSV"""
        try:
            if format.lower() == "pdf":
                return await self._export_pdf(report_type, report_params)
            elif format.lower() == "excel":
                return await self._export_excel(report_type, report_params)
            elif format.lower() == "csv":
                return await self._export_csv(report_type, report_params)
            else:
                raise ValueError(f"Unsupported format: {format}")
        except Exception as e:
            logger.error("Report export failed", report_type=report_type, format=format, error=str(e), exc_info=True)
            raise
    
    async def _export_pdf(self, report_type: str, report_params: Dict[str, Any]) -> Dict[str, Any]:
        """Export report to PDF using ReportLab"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
        )
        elements.append(Paragraph(f"{report_type.replace('_', ' ').title()} Report", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Get report data based on type
        report_data = await self._get_report_data(report_type, report_params)
        
        if not report_data or "data" not in report_data:
            elements.append(Paragraph("No data available for this report.", styles['Normal']))
        else:
            # Create table from data
            data = report_data["data"]
            if isinstance(data, list) and len(data) > 0:
                # Headers
                if isinstance(data[0], dict):
                    headers = list(data[0].keys())
                    table_data = [headers]
                    
                    # Data rows
                    for row in data:
                        table_data.append([str(row.get(h, "")) for h in headers])
                else:
                    table_data = data
                
                # Create table
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                ]))
                elements.append(table)
            else:
                elements.append(Paragraph("No data rows available.", styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Return PDF as base64 or save to file
        pdf_bytes = buffer.getvalue()
        
        # Save to temporary file
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            tmp_file.write(pdf_bytes)
            tmp_path = tmp_file.name
        
        return {
            "report_type": report_type,
            "format": "pdf",
            "file_path": tmp_path,
            "file_size": len(pdf_bytes),
            "message": "PDF report generated successfully"
        }
    
    async def _export_excel(self, report_type: str, report_params: Dict[str, Any]) -> Dict[str, Any]:
        """Export report to Excel using openpyxl"""
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.replace('_', ' ').title()
        
        # Get report data
        report_data = await self._get_report_data(report_type, report_params)
        
        if not report_data or "data" not in report_data:
            ws['A1'] = "No data available for this report."
        else:
            data = report_data["data"]
            if isinstance(data, list) and len(data) > 0:
                # Headers
                if isinstance(data[0], dict):
                    headers = list(data[0].keys())
                    ws.append(headers)
                    
                    # Style header row
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Data rows
                    for row in data:
                        ws.append([str(row.get(h, "")) for h in headers])
                else:
                    # If data is already a list of lists
                    for row in data:
                        ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_path = tmp_file.name
        
        file_size = os.path.getsize(tmp_path)
        
        return {
            "report_type": report_type,
            "format": "excel",
            "file_path": tmp_path,
            "file_size": file_size,
            "message": "Excel report generated successfully"
        }
    
    async def _export_csv(self, report_type: str, report_params: Dict[str, Any]) -> Dict[str, Any]:
        """Export report to CSV"""
        import csv
        
        # Get report data
        report_data = await self._get_report_data(report_type, report_params)
        
        if not report_data or "data" not in report_data:
            csv_data = "No data available for this report."
        else:
            data = report_data["data"]
            if isinstance(data, list) and len(data) > 0:
                output = io.StringIO()
                if isinstance(data[0], dict):
                    headers = list(data[0].keys())
                    writer = csv.DictWriter(output, fieldnames=headers)
                    writer.writeheader()
                    writer.writerows(data)
                else:
                    writer = csv.writer(output)
                    writer.writerows(data)
                csv_data = output.getvalue()
            else:
                csv_data = "No data rows available."
        
        # Save to temporary file
        import tempfile
        with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', newline='') as tmp_file:
            tmp_file.write(csv_data)
            tmp_path = tmp_file.name
        
        file_size = os.path.getsize(tmp_path)
        
        return {
            "report_type": report_type,
            "format": "csv",
            "file_path": tmp_path,
            "file_size": file_size,
            "message": "CSV report generated successfully"
        }
    
    async def generate_production_turnaround_report(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> Dict[str, Any]:
        """Generate production turnaround time report"""
        from backend.models.production_order import ProductionOrder, ProductionOrderStatus
        from sqlalchemy import select, and_, func
        from datetime import datetime, timezone
        
        query = select(ProductionOrder)
        
        if start_date:
            query = query.where(ProductionOrder.created_at >= datetime.combine(start_date, datetime.min.time()).replace(tzinfo=timezone.utc))
        if end_date:
            query = query.where(ProductionOrder.created_at <= datetime.combine(end_date, datetime.max.time()).replace(tzinfo=timezone.utc))
        
        query = query.where(
            ProductionOrder.status.in_([
                ProductionOrderStatus.COMPLETED,
                ProductionOrderStatus.DELIVERED
            ])
        )
        
        result = await self.db.execute(query)
        orders = result.scalars().all()
        
        turnaround_times = []
        for order in orders:
            if order.completed_at and order.created_at:
                turnaround = (order.completed_at - order.created_at).total_seconds() / 3600  # Hours
                turnaround_times.append({
                    "po_number": order.po_number,
                    "client_name": order.client_name,
                    "created_at": order.created_at.isoformat(),
                    "completed_at": order.completed_at.isoformat(),
                    "turnaround_hours": round(turnaround, 2),
                    "turnaround_days": round(turnaround / 24, 2)
                })
        
        avg_turnaround = sum(t["turnaround_hours"] for t in turnaround_times) / len(turnaround_times) if turnaround_times else 0
        
        return {
            "report_type": "production_turnaround",
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
            "total_orders": len(turnaround_times),
            "average_turnaround_hours": round(avg_turnaround, 2),
            "average_turnaround_days": round(avg_turnaround / 24, 2),
            "orders": turnaround_times
        }
    
    async def generate_production_workload_report(
        self,
        user_id: Optional[UUID] = None
    ) -> Dict[str, Any]:
        """Generate production workload report by user"""
        from backend.models.production_assignment import ProductionAssignment, AssignmentStatus
        from backend.models.production_order import ProductionOrder, ProductionOrderStatus
        from sqlalchemy import select, and_, func
        from backend.models.user import User
        
        query = select(
            ProductionAssignment.user_id,
            func.count(ProductionAssignment.id).label('total_assignments'),
            func.count(
                func.case(
                    (ProductionAssignment.status == AssignmentStatus.IN_PROGRESS, 1),
                    else_=None
                )
            ).label('in_progress'),
            func.count(
                func.case(
                    (ProductionAssignment.status == AssignmentStatus.COMPLETED, 1),
                    else_=None
                )
            ).label('completed')
        ).group_by(ProductionAssignment.user_id)
        
        if user_id:
            query = query.where(ProductionAssignment.user_id == user_id)
        
        result = await self.db.execute(query)
        workload_data = result.all()
        
        workload_report = []
        for row in workload_data:
            user_result = await self.db.execute(
                select(User).where(User.id == row.user_id)
            )
            user = user_result.scalar_one_or_none()
            
            workload_report.append({
                "user_id": row.user_id,
                "username": user.username if user else "Unknown",
                "role": user.role if user else "Unknown",
                "total_assignments": row.total_assignments,
                "in_progress": row.in_progress,
                "completed": row.completed,
                "pending": row.total_assignments - row.in_progress - row.completed
            })
        
        return {
            "report_type": "production_workload",
            "workload_by_user": workload_report
        }
    
    async def generate_missed_deadlines_report(
        self,
        days_overdue: int = 0
    ) -> Dict[str, Any]:
        """Generate report of production orders with missed deadlines"""
        from backend.models.production_order import ProductionOrder, ProductionOrderStatus
        from sqlalchemy import select, and_
        from datetime import datetime, timezone, timedelta
        
        cutoff_date = datetime.now(timezone.utc) - timedelta(days=days_overdue)
        
        result = await self.db.execute(
            select(ProductionOrder)
            .where(
                and_(
                    ProductionOrder.deadline.isnot(None),
                    ProductionOrder.deadline < cutoff_date,
                    ProductionOrder.status.notin_([
                        ProductionOrderStatus.COMPLETED,
                        ProductionOrderStatus.DELIVERED,
                        ProductionOrderStatus.CANCELLED
                    ])
                )
            )
            .order_by(ProductionOrder.deadline.asc())
        )
        overdue_orders = result.scalars().all()
        
        overdue_list = []
        for order in overdue_orders:
            days_overdue_count = (datetime.now(timezone.utc) - order.deadline).days if order.deadline else 0
            overdue_list.append({
                "po_number": order.po_number,
                "client_name": order.client_name,
                "deadline": order.deadline.isoformat() if order.deadline else None,
                "days_overdue": days_overdue_count,
                "status": order.status.value if hasattr(order.status, 'value') else str(order.status)
            })
        
        return {
            "report_type": "missed_deadlines",
            "total_overdue": len(overdue_list),
            "orders": overdue_list
        }
    
    async def _get_report_data(self, report_type: str, report_params: Dict[str, Any]) -> Dict[str, Any]:
        """Get report data based on type"""
        # This is a helper method to fetch data for different report types
        # It calls the appropriate report generation method
        
        if report_type == "revenue_summary":
            start_date = datetime.fromisoformat(report_params.get("start_date", date.today().isoformat())).date()
            end_date = datetime.fromisoformat(report_params.get("end_date", date.today().isoformat())).date()
            return await self.generate_revenue_summary(start_date, end_date)
        elif report_type == "daily_log":
            log_date = datetime.fromisoformat(report_params.get("log_date", date.today().isoformat())).date()
            return await self.generate_daily_log_report(log_date)
        elif report_type == "makegood":
            start_date = datetime.fromisoformat(report_params.get("start_date", date.today().isoformat())).date() if report_params.get("start_date") else None
            end_date = datetime.fromisoformat(report_params.get("end_date", date.today().isoformat())).date() if report_params.get("end_date") else None
            return await self.generate_makegood_report(start_date, end_date)
        elif report_type == "audio_activity":
            start_date = datetime.fromisoformat(report_params.get("start_date", date.today().isoformat())).date()
            end_date = datetime.fromisoformat(report_params.get("end_date", date.today().isoformat())).date()
            return await self.generate_audio_activity_report(start_date, end_date)
        elif report_type == "cut_rotation_performance":
            copy_id = report_params.get("copy_id")
            return await self.generate_cut_rotation_performance(copy_id=copy_id)
        elif report_type == "live_read_performance":
            start_date = datetime.fromisoformat(report_params.get("start_date", date.today().isoformat())).date()
            end_date = datetime.fromisoformat(report_params.get("end_date", date.today().isoformat())).date()
            return await self.generate_live_read_performance(start_date, end_date)
        elif report_type == "fcc_compliance":
            start_date = datetime.fromisoformat(report_params.get("start_date", date.today().isoformat())).date()
            end_date = datetime.fromisoformat(report_params.get("end_date", date.today().isoformat())).date()
            return await self.generate_fcc_compliance_log(start_date, end_date)
        elif report_type == "production_turnaround":
            start_date = datetime.fromisoformat(report_params.get("start_date")).date() if report_params.get("start_date") else None
            end_date = datetime.fromisoformat(report_params.get("end_date")).date() if report_params.get("end_date") else None
            return await self.generate_production_turnaround_report(start_date, end_date)
        elif report_type == "production_workload":
            user_id = report_params.get("user_id")
            return await self.generate_production_workload_report(user_id)
        elif report_type == "missed_deadlines":
            days_overdue = report_params.get("days_overdue", 0)
            return await self.generate_missed_deadlines_report(days_overdue)
        else:
            # Default: return empty data structure
            return {"data": []}

